"""
Diagnostic script: checks which Excel country names fail to match country_codes.json.
Replicates the exact matching logic from index.html.
"""

import json
import openpyxl
import os

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE, "GTF Coaches Data.xlsx")
JSON_PATH  = os.path.join(BASE, "country_codes.json")

# ── 1. Load country_codes.json ────────────────────────────────────────────────
with open(JSON_PATH, encoding="utf-8") as f:
    codes_json = json.load(f)

countries_dict = codes_json.get("countries", {})
iso_to_label = {}
for iso3, obj in countries_dict.items():
    if obj and isinstance(obj.get("label"), str) and obj["label"].strip():
        iso_to_label[iso3] = obj["label"].strip()

def normalise(s):
    return str(s or "").strip().lower()

norm_label_to_label = {}
for label in iso_to_label.values():
    norm_label_to_label[normalise(label)] = label

# ── 2. Manual aliases (copied verbatim from index.html) ───────────────────────
COUNTRY_ALIASES = {
    "congo, the democratic republic of the": "Congo, Dem. Rep. of the",
    "kyrgyzstan":    "Kyrgyz Republic",
    "palestine":     "West Bank and Gaza",
    "south sudan":   "South Sudan, Republic of",
    "turkey":        "Türkiye",
    "united states": "United States of America",
}

# ── 3. Load Excel ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

# Read headers from first row
headers = []
for cell in ws[1]:
    headers.append(str(cell.value) if cell.value is not None else "")

print("Excel columns found:", headers)
print()

# Detect country column (same logic as JS)
COUNTRY_KEYS = ["country", "nation", "pais", "país"]
COUNT_KEYS   = ["grand total", "total", "coaches", "coach count", "number of coaches"]

country_col_idx = None
for i, h in enumerate(headers):
    if normalise(h) in COUNTRY_KEYS:
        country_col_idx = i
        break
if country_col_idx is None:
    country_col_idx = 0

# Coach count column: skip country column, look for COUNT_KEYS substring
coach_col_idx = None
for i, h in enumerate(headers):
    if i == country_col_idx:
        continue
    if any(c in normalise(h) for c in COUNT_KEYS):
        coach_col_idx = i
        break
if coach_col_idx is None:
    coach_col_idx = len(headers) - 1  # fallback: last column

print(f"Country column : [{country_col_idx}] '{headers[country_col_idx]}'")
print(f"Coach count col: [{coach_col_idx}] '{headers[coach_col_idx]}'")
print()

# ── 4. Aggregate raw data ─────────────────────────────────────────────────────
raw_data = {}          # normKey -> count
skipped_zero = []      # (raw_country, count)

for row in ws.iter_rows(min_row=2, values_only=True):
    raw_country = str(row[country_col_idx] or "").strip()
    raw_count_val = row[coach_col_idx]
    try:
        raw_count = float(raw_count_val or 0)
    except (TypeError, ValueError):
        raw_count = 0

    if not raw_country or raw_count == 0:
        if raw_country:
            skipped_zero.append((raw_country, raw_count))
        continue

    norm_key = normalise(raw_country)
    raw_data[norm_key] = raw_data.get(norm_key, 0) + raw_count

# ── 5. Simulate matching ──────────────────────────────────────────────────────
matched_alias   = []   # (normKey, count, canonLabel)
matched_exact   = []
matched_partial = []
unmatched       = []   # (normKey, count)

for norm_key, count in sorted(raw_data.items(), key=lambda x: -x[1]):
    # 1) Manual alias
    alias_label = COUNTRY_ALIASES.get(norm_key)
    canon_label = norm_label_to_label.get(normalise(alias_label)) if alias_label else None

    if canon_label:
        matched_alias.append((norm_key, count, canon_label))
        continue

    # 2) Exact normalised match
    canon_label = norm_label_to_label.get(norm_key)
    if canon_label:
        matched_exact.append((norm_key, count, canon_label))
        continue

    # 3) Partial match
    found = None
    for nl, cl in norm_label_to_label.items():
        if nl in norm_key or norm_key in nl:
            found = cl
            break
    if found:
        matched_partial.append((norm_key, count, found))
        continue

    unmatched.append((norm_key, count))

# ── 6. Report ─────────────────────────────────────────────────────────────────
SEP = "-" * 72

print(SEP)
print(f"  ALIAS MATCHES ({len(matched_alias)} countries)")
print(SEP)
for norm_key, count, canon in matched_alias:
    print(f"  {int(count):>5}  '{norm_key}'  ->  '{canon}'")

print()
print(SEP)
print(f"  EXACT MATCHES ({len(matched_exact)} countries)")
print(SEP)
for norm_key, count, canon in matched_exact:
    print(f"  {int(count):>5}  '{norm_key}'  ->  '{canon}'")

print()
print(SEP)
print(f"  PARTIAL MATCHES ({len(matched_partial)} countries)")
print(SEP)
for norm_key, count, canon in matched_partial:
    print(f"  {int(count):>5}  '{norm_key}'  ->  '{canon}'")

print()
print(SEP)
print(f"  SKIPPED (zero/blank coach count) — {len(skipped_zero)} rows")
print(SEP)
for raw_country, count in sorted(skipped_zero, key=lambda x: x[0]):
    print(f"         '{raw_country}'  (count={count})")

print()
print(SEP)
if unmatched:
    print(f"  *** UNMATCHED — DATA LOSS ({len(unmatched)} countries) ***")
    print(SEP)
    total_lost = 0
    for norm_key, count in sorted(unmatched, key=lambda x: -x[1]):
        print(f"  {int(count):>5}  '{norm_key}'")
        total_lost += count
    print()
    print(f"  Total coaches lost: {int(total_lost)}")
else:
    print("  ALL COUNTRIES MATCHED SUCCESSFULLY — no data loss.")
print(SEP)

total_matched = sum(c for _, c, _ in matched_alias) \
              + sum(c for _, c, _ in matched_exact) \
              + sum(c for _, c, _ in matched_partial)
total_unmatched = sum(c for _, c in unmatched)

print()
print("SUMMARY")
print(f"  Alias matches   : {len(matched_alias)} countries")
print(f"  Exact matches   : {len(matched_exact)} countries")
print(f"  Partial matches : {len(matched_partial)} countries")
print(f"  Unmatched       : {len(unmatched)} countries  ({int(total_unmatched)} coaches lost)")
print(f"  Skipped (zero)  : {len(skipped_zero)} rows")
print(f"  Total coaches matched : {int(total_matched)}")
print(f"  Total coaches lost    : {int(total_unmatched)}")
